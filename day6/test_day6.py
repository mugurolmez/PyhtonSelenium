from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep #chrome kapanmasın dıye
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path #klasör acmak için
from datetime import date
import os
import openpyxl
from constants import globalConstants
#prefix=> ön ek test_
#postfix


class  Test_DemoClass2:
    #her testten once çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)#exist ok ilgili klasor olusturulmussa tekrar olusturma demek
        
        #günün tarihini al bu tarihle bir klasor var mı kontrol et yoksa olustur
   
       
        
    #her testten sonra çağırılır
    def teardown_method(self):
        self.driver.quit()

    def readData(self):
        print("x")
    #3A Act arrange Asssert
    #setup->test_demoFunc->teardown
    def test_demoFunc(self):
        text ="Hello"
        assert True
    #setup->test_demoFunc->teardown
    def test_demo2(self):
        assert True #ilgili testin bağlı oldugu sonuc true mi false mi test basarılı basarısızmı 

    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("day6/data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
    #coklu kullanıcı adı şifre deneme anatasyonu
    @pytest.mark.parametrize("username,password",getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
       
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        
        loginBtn=self.driver.find_element(By.NAME,"login-button")
        
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

        


    