from datetime import date
import inspect
from pathlib import Path
import openpyxl
import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep #chrome kapanmasın dıye
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import os
from constants import globalConstants

class Test_homework6:
    def my_function():
        pass
    def screen_shot(self):
        test_name = inspect.currentframe().f_back.f_code.co_name#bir üst fonksiyonun ismimi almak için 
        self.driver.save_screenshot(f"{self.folderPath}/{test_name}.png")
    
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath =os.path.join(os.path.dirname(os.path.abspath(__file__)), str(date.today()))#bulkundugum klasorun yolu
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()
    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("day6homework2/data/valid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]

        totalRows = selectedSheet.max_row
        data=[]
        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
    def waitForElementVisible(self,locator,timeout=5):
       WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

    def loginButtonClick(self):
        self.waitForElementVisible((By.ID,globalConstants.loginButton))
        loginButton = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginButton.click()

    def validLogin(self):
        self.inputName(globalConstants.standardUser)
        self.inputPassword(globalConstants.secretSauce)
        self.loginButtonClick()

    def loginuser(self,username,password):
        self.inputName(username)
        self.inputPassword(password)
        self.loginButtonClick()

    def add_all_to_cart_click(self):
        self.validLogin()
        list=(globalConstants.cartList)
        i=0
        for i in range(len(list)):

            self.waitForElementVisible((By.ID,list[i]))
            add_cart=self.driver.find_element(By.ID,list[i])
            add_cart.click()
            
    def inputName(self,userName):
        self.waitForElementVisible((By.ID,globalConstants.userName))
        inputName = self.driver.find_element(By.ID,globalConstants.userName)
        inputName.send_keys(userName)

    def inputPassword(self,password):
        self.waitForElementVisible((By.ID,globalConstants.password))      
        inputPassword=self.driver.find_element(By.ID,globalConstants.password)
        inputPassword.send_keys(password)
            
    def xButtonClick(self):
        xbutton = self.driver.find_element(By.CLASS_NAME, globalConstants.errorButton)
        xbutton.click()
        

#------------------------------------------------------------

    def  test_nullnameAndPassword(self):
        self.loginButtonClick()   

        self.waitForElementVisible((globalConstants.errorMessageLocator))
        errorMessage = self.driver.find_element(By.XPATH,globalConstants.errorMessageValue)
        assert errorMessage.text == globalConstants.nullNameAndPasswoerdErrorMessage
        self.screen_shot()
        
        


    def test_nullPassword(self):
        
        self.inputName(globalConstants.standardUser)
        self.loginButtonClick()

      
        self.waitForElementVisible((globalConstants.errorMessageLocator))
        errorMessage=self.driver.find_element(By.XPATH,globalConstants.errorMessageValue)
        
        assert errorMessage.text == globalConstants.nullPasswordErrorMessage
        self.screen_shot()
        


    def test_lockedUser(self):
        self.inputName(globalConstants.lockedUsername)
       
        self.inputPassword(globalConstants.secretSauce)      
    
        self.loginButtonClick()
      

        self.waitForElementVisible((globalConstants.errorMessageLocator))
        errorMessage=self.driver.find_element(By.XPATH,globalConstants.errorMessageValue)
        assert errorMessage.text == globalConstants.lockedUserErrorMessage
        self.screen_shot()


    def test_xButton(self):
        self.loginButtonClick() 
        self.xButtonClick()
        buttonIsVisible=len(self.driver.find_elements(By.CLASS_NAME, globalConstants.errorButton))#elements liste olacak yoksa len çalışmaz
                
        assert buttonIsVisible==0#butonun gorunmez olduguğunu dogrulamak ıcın assert not xbutton.disp false geldiğinde dogru olacak
        self.screen_shot()
    
    def test_loginUser(self):
        
        self.validLogin()
        ifPageLoaded=(self.driver.current_url==globalConstants.inventoryURL)
        
        list = self.driver.find_elements(By.CLASS_NAME, globalConstants.inventoryItem)
        #açılan sayfadaki inventory_itemleri listele
        assert(ifPageLoaded and len(list)==6)
        self.screen_shot()

    def test_add_to_cart_click(self):
        self.validLogin()

        saucelabsBackpack=self.driver.find_element(By.CLASS_NAME,globalConstants.inventoryItemName)
        saucelabsBackpack.click()

        ifPageLoaded=self.driver.current_url==globalConstants.labsBackPackURL
        assert (ifPageLoaded==True)
        self.screen_shot()


    def test_add_all_to_cart(self):
        self.add_all_to_cart_click()

        self.waitForElementVisible(globalConstants.shopingCartLinkLocator)
        cartLink=self.driver.find_element(By.CLASS_NAME,globalConstants.shopingCartLinkValue)
        cartLink.click()
    
        self.waitForElementVisible(globalConstants.inventoryContainerLocator)
        list=self.driver.find_elements(By.CLASS_NAME,globalConstants.cartItem)

        assert (len(list)==6)
        self.screen_shot()


    @pytest.mark.parametrize("username,password",getData())
    def test_valid_users(self,username,password):
        self.inputName(username)
        self.inputPassword(password)
        self.loginButtonClick()

        ifPageLoaded=self.driver.current_url==globalConstants.inventoryURL
        assert (ifPageLoaded==True)
        self.screen_shot()

    def test_backtoproducts(self):
       
        self.validLogin()
        self.waitForElementVisible(globalConstants.sauceLabsBackpacklocator)
        labsBackpack = self.driver.find_element(By.XPATH,globalConstants.sauceLabsBackpackXpath)
        labsBackpack.click()
      

        self.waitForElementVisible(globalConstants.backToProtuctLocator)
        backtoproducts =self.driver.find_element(By.CSS_SELECTOR, globalConstants.backtoProductcsselector)
        backtoproducts.click()
        
        
        self.waitForElementVisible(globalConstants.productsTittleLocator)
       
    
        assert self.driver.find_element(By.XPATH, globalConstants.productsTitleXpath).text == "Products"




        




    
        
        
   
    



    


    